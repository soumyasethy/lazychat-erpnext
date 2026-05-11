"""Knowledge Base — search attached files in a Lazychat Knowledge Base
doctype row.

MVP scope (slice 1):
  * Multi-format text extraction: txt/md/csv/json/yaml (UTF-8) + xlsx
    (openpyxl) + pdf (pdfplumber → pypdf fallback) + docx (python-docx).
    Each format guarded by try/import so a missing optional dep degrades
    to a clear error instead of a 500.
  * Keyword paragraph search — splits extracted text into paragraphs,
    filters those containing ALL query terms (case-insensitive), returns
    top N with file metadata. Cheap, transparent, no external services.
  * No persistent index, no embeddings — every search re-extracts. Fine
    for KBs under ~50 files. Embeddings + vector search are slice 2.

Doctype is `Lazychat Knowledge Base`. Files are linked via the standard
Frappe File doctype with attached_to_doctype="Lazychat Knowledge Base"
and attached_to_name=<kb_name>. Users attach files via the Desk's
sidebar Attachments UI on the parent KB doc.
"""

import io
import re

import frappe

KB_DOCTYPE = "Lazychat Knowledge Base"

_TEXT_EXTENSIONS = (
	".txt", ".md", ".markdown", ".csv", ".tsv", ".json", ".yaml", ".yml",
	".log", ".html", ".htm", ".xml", ".sql", ".py", ".js", ".ts",
)


def _visible_kb_filter(user=None):
	"""Filter the user can apply to Lazychat Knowledge Base queries: enabled +
	(public OR owned by them). Returns a (filters, or_filters) tuple ready for
	frappe.get_all."""
	user = user or frappe.session.user
	return (
		{"enabled": 1},
		[{"is_public": 1}, {"owner": user}],
	)


def list_kbs_for_user(user=None):
	"""Return enabled KBs the user can see, with file counts."""
	filters, or_filters = _visible_kb_filter(user=user)
	rows = frappe.get_all(
		KB_DOCTYPE,
		filters=filters,
		or_filters=or_filters,
		fields=["name", "title", "description", "is_public", "owner"],
		order_by="is_public desc, title asc",
	)
	out = []
	for r in rows:
		try:
			file_count = frappe.db.count(
				"File",
				{"attached_to_doctype": KB_DOCTYPE, "attached_to_name": r["name"]},
			)
		except Exception:
			file_count = 0
		out.append(
			{
				"name": r["name"],
				"title": r.get("title") or r["name"],
				"description": r.get("description") or "",
				"is_public": bool(r.get("is_public")),
				"owner_user": r.get("owner"),
				"file_count": file_count,
			}
		)
	return out


def _user_can_read_kb(kb_name, user=None):
	user = user or frappe.session.user
	row = frappe.db.get_value(
		KB_DOCTYPE, kb_name, ["is_public", "owner", "enabled"], as_dict=True
	)
	if not row or not row.enabled:
		return False
	return bool(row.is_public) or row.owner == user


def get_kb_files(kb_name):
	if not _user_can_read_kb(kb_name):
		return {"error": f"knowledge base not found or not accessible: {kb_name}"}
	rows = frappe.get_all(
		"File",
		filters={"attached_to_doctype": KB_DOCTYPE, "attached_to_name": kb_name},
		fields=["name", "file_name", "file_url", "is_private", "file_size", "file_type"],
		order_by="file_name asc",
	)
	# Tier H2 — surface per-file embedding status so the agent (and the user)
	# knows whether semantic search is available or we're still on keyword-only.
	try:
		from lazychat_mcp_erpnext.desk_assistant import embeddings as _emb

		index_status = _emb.kb_index_status(kb_name)
	except Exception:
		index_status = {}
	for r in rows:
		st = index_status.get(r["name"]) or {}
		r["embedding_status"] = st.get("status") or "pending"
		r["chunk_count"] = st.get("chunk_count") or 0
		r["embedded_count"] = st.get("embedded_count") or 0
	return {"ok": True, "kb_name": kb_name, "count": len(rows), "files": rows}


# ---------- multi-format text extraction ----------

def _extract_text_xlsx(content_bytes):
	from openpyxl import load_workbook  # type: ignore

	wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
	parts = []
	for sheet in wb.sheetnames:
		parts.append(f"## Sheet: {sheet}")
		ws = wb[sheet]
		for row in ws.iter_rows(values_only=True):
			cells = ["" if v is None else str(v) for v in row]
			parts.append(" | ".join(cells))
		parts.append("")
	return "\n".join(parts)


def _extract_text_pdf(content_bytes):
	try:
		import pdfplumber  # type: ignore

		with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
			return "\n\n".join((p.extract_text() or "") for p in pdf.pages)
	except ImportError:
		pass
	try:
		from pypdf import PdfReader  # type: ignore

		reader = PdfReader(io.BytesIO(content_bytes))
		return "\n\n".join((p.extract_text() or "") for p in reader.pages)
	except ImportError:
		raise RuntimeError(
			"PDF extraction needs pdfplumber or pypdf installed in the bench. "
			"Run: bench pip install pdfplumber"
		)


def _extract_text_docx(content_bytes):
	try:
		from docx import Document as DocxDocument  # type: ignore
	except ImportError:
		raise RuntimeError(
			"DOCX extraction needs python-docx installed in the bench. "
			"Run: bench pip install python-docx"
		)
	doc = DocxDocument(io.BytesIO(content_bytes))
	return "\n".join(p.text for p in doc.paragraphs if p.text)


def extract_file_text(file_doc, max_chars=20000):
	"""Return (text, format_used, error_message_or_None) for a Frappe File doc.

	text=None means extraction failed; the error_message describes why.
	max_chars is enforced AFTER extraction, so very large files don't
	balloon the LLM context window.
	"""
	fname = (file_doc.file_name or "").lower()
	try:
		raw = file_doc.get_content()
	except FileNotFoundError:
		return None, None, "underlying file is missing on disk"

	if isinstance(raw, str):
		return raw[:max_chars], "text", None

	if not isinstance(raw, bytes):
		return str(raw)[:max_chars], "unknown", None

	# Try UTF-8 first; covers explicit text extensions plus anything with
	# accidental utf-8 content.
	if any(fname.endswith(ext) for ext in _TEXT_EXTENSIONS):
		try:
			return raw.decode("utf-8", errors="replace")[:max_chars], "text", None
		except Exception as e:
			return None, None, f"text-decode-error: {e}"

	if fname.endswith(".pdf"):
		try:
			return _extract_text_pdf(raw)[:max_chars], "pdf", None
		except Exception as e:
			return None, None, f"pdf-error: {e}"

	if fname.endswith(".xlsx") or fname.endswith(".xlsm"):
		try:
			return _extract_text_xlsx(raw)[:max_chars], "xlsx", None
		except Exception as e:
			return None, None, f"xlsx-error: {e}"

	if fname.endswith(".docx"):
		try:
			return _extract_text_docx(raw)[:max_chars], "docx", None
		except Exception as e:
			return None, None, f"docx-error: {e}"

	# Fallback — try utf-8 with replacement
	try:
		return raw.decode("utf-8")[:max_chars], "text-fallback", None
	except UnicodeDecodeError:
		return None, None, "binary file — extension not in the supported list (.txt .md .csv .json .yaml .pdf .xlsx .docx)"


# ---------- keyword search ----------

_PARA_SPLIT = re.compile(r"\n\s*\n+")
_TOKEN_RE = re.compile(r"[A-Za-z0-9_]{2,}")


def _matches_all_terms(paragraph, terms):
	low = paragraph.lower()
	return all(t in low for t in terms)


def _make_snippet(paragraph, terms, max_len=400):
	# Pick the substring centred on the first term hit.
	low = paragraph.lower()
	idx = -1
	for t in terms:
		idx = low.find(t)
		if idx >= 0:
			break
	if idx < 0:
		return paragraph.strip()[:max_len]
	start = max(0, idx - max_len // 4)
	end = min(len(paragraph), start + max_len)
	prefix = "…" if start > 0 else ""
	suffix = "…" if end < len(paragraph) else ""
	return prefix + paragraph[start:end].strip() + suffix


def search(query, kb_name=None, max_chunks=8, max_chars_per_file=30000):
	"""Search across one named KB or all visible ones.

	Returns:
	  {ok, query, kb_names_searched, files_scanned, chunks: [
	    {kb_name, file_name, file_url, snippet, format, score?, vector_rank?, keyword_rank?}
	  ]}

	Routing (Tier H2):
	  - If any KB Chunk rows exist for the scope → delegate to embeddings.hybrid_search
	    (cosine-over-embeddings + keyword + RRF fusion). When chunks have no
	    embeddings yet, falls back internally to keyword-only.
	  - Else → file-walk + paragraph match (original keyword-only path below).
	    This path is the v1 fallback for KBs that haven't been indexed yet.
	"""
	terms = [t.lower() for t in _TOKEN_RE.findall(query)]
	# Drop stop-word-ish single chars and dedupe
	terms = [t for t in dict.fromkeys(terms) if len(t) >= 2]
	if not terms:
		return {"error": "query needs at least one alphanumeric term of length >= 2"}

	# Resolve which KBs to scan
	if kb_name:
		if not _user_can_read_kb(kb_name):
			return {"error": f"knowledge base not found or not accessible: {kb_name}"}
		kb_names = [kb_name]
	else:
		visible = list_kbs_for_user()
		kb_names = [kb["name"] for kb in visible]
		if not kb_names:
			return {"ok": True, "query": query, "kb_names_searched": [], "files_scanned": 0, "chunks": []}

	# H2 path: prefer hybrid search when any KB Chunk rows exist for this scope.
	try:
		any_chunks = frappe.db.count(
			"Lazychat KB Chunk", {"parent_kb": ["in", kb_names]}
		)
	except Exception:
		any_chunks = 0
	if any_chunks > 0:
		try:
			from lazychat_mcp_erpnext.desk_assistant import embeddings as _emb

			chunks = _emb.hybrid_search(query, kb_names, max_chunks=max_chunks)
			return {
				"ok": True,
				"query": query,
				"kb_names_searched": kb_names,
				"files_scanned": len({c["file_name"] for c in chunks}),
				"chunks": chunks,
				"_note": (
					f"Hybrid retrieval (cosine + keyword RRF) over {any_chunks} indexed chunks. "
					f"Returned {len(chunks)} of max {max_chunks}."
				),
			}
		except Exception as e:
			# Fall through to legacy keyword-only path on any failure
			frappe.log_error(frappe.get_traceback(), f"lazychat hybrid_search failed: {e}")

	files = frappe.get_all(
		"File",
		filters={
			"attached_to_doctype": KB_DOCTYPE,
			"attached_to_name": ["in", kb_names],
		},
		fields=["name", "file_name", "file_url", "attached_to_name"],
		order_by="modified desc",
	)

	chunks = []
	files_scanned = 0
	for f in files:
		try:
			file_doc = frappe.get_doc("File", f["name"])
		except Exception:
			continue
		files_scanned += 1
		text, fmt, err = extract_file_text(file_doc, max_chars=max_chars_per_file)
		if not text:
			# Still surface unreadable files so the user can fix them.
			if err:
				chunks.append(
					{
						"kb_name": f["attached_to_name"],
						"file_name": f["file_name"],
						"file_url": f["file_url"],
						"snippet": f"[unreadable: {err}]",
						"format": "error",
					}
				)
			continue
		paragraphs = [p for p in _PARA_SPLIT.split(text) if p.strip()]
		for p in paragraphs:
			if _matches_all_terms(p, terms):
				chunks.append(
					{
						"kb_name": f["attached_to_name"],
						"file_name": f["file_name"],
						"file_url": f["file_url"],
						"snippet": _make_snippet(p, terms),
						"format": fmt,
					}
				)
				if len(chunks) >= max_chunks:
					break
		if len(chunks) >= max_chunks:
			break

	return {
		"ok": True,
		"query": query,
		"kb_names_searched": kb_names,
		"files_scanned": files_scanned,
		"chunks": chunks,
		"_note": (
			f"Keyword paragraph search over {files_scanned} file(s) across "
			f"{len(kb_names)} KB(s). Returned {len(chunks)} of "
			f"max {max_chunks} chunks. Vector / embedding search is on the roadmap."
		),
	}
